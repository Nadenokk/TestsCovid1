# -*- coding: utf-8 -*-
import glob

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
import logging, os
from openpyxl import load_workbook

class OtchetIssledovanie(unittest.TestCase):
    def setUp(self):
        download_dir = "C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel"
        chrome_options = webdriver.ChromeOptions()
        preferences = {"download.default_directory": download_dir,
                       "directory_upgrade": True,
                       "safebrowsing.enabled": True}
        chrome_options.add_experimental_option("prefs", preferences)
        self.driver = webdriver.Chrome("C:/Users/user/Downloads/chromedriver.exe", options=chrome_options)
        self.driver.set_window_size(1024, 600)
        self.driver.maximize_window()
        self.driver.implicitly_wait(60)
        self.verificationErrors = []
        self.accept_next_alert = True

        '''
        options = webdriver.ChromeOptions()
        #options.add_argument()
        pathdow ='C:/Users/user/PycharmProjects/TestsCovid1/COVID_19_9880/Filling_Lab_Tablets/downloads_exel'



        #self.driver = webdriver.Chrome(chrome_options=options)
        self.driver.set_window_size(1024, 600)
        self.driver.maximize_window()
        self.driver.implicitly_wait(60)
        self.verificationErrors = []
        self.accept_next_alert = True
        '''

        '''    
        def genlog(self):    
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        if not os.path.exists("Logs"):
            os.mkdir("Logs")
        handler = logging.FileHandler(str('logs/' + (time.strftime(''%d.%m.%Y_%H.%M_'', (time.localtime())))  + 'lab_planshet_1.log'))
        handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        try:
            self.lab_planshet()
        except Exception as e:
            logger.error('Error detected', exc_info=True)
        else:
            logger.info('Test complete without errors')    

    def test_lab_planshet(self):
        self.genlog()
    '''

    def test_otchet_issledovanie(self):
        driver = self.driver
        driver.get("http://auraep.ru:9880/business/dashboard/dashboard.xhtml#")
        # driver.get("https://rpn19.ru:11443/documents/")
        driver.find_element_by_id("form:usernameInput").click()
        driver.find_element_by_id("form:usernameInput").clear()
        driver.find_element_by_id("form:usernameInput").send_keys("supervisor")
        driver.find_element_by_id("form:passwordInput").click()
        driver.find_element_by_id("form:passwordInput").clear()
        driver.find_element_by_id("form:passwordInput").send_keys("Ivwdk1Rp")
        driver.find_element_by_css_selector("span.ui-button-text.ui-c").click()
        driver.find_element_by_css_selector(
            "#j_idt68 > div.nano.layout-tabmenu-nav > ul > li:nth-child(13) > a > div").click()
        driver.find_element_by_id("reportsForm:j_idt77:11:j_idt79").click()
        driver.find_element_by_id("buildForm:sendInstitutions").click()
        time.sleep(7)
        #driver.find_element_by_id("buildForm:sendInstitutions_label").click()
        driver.find_element_by_css_selector("#buildForm\:sendInstitutions_panel").click()
        #driver.find_element_by_xpath("//label[@value='Администрация Адмиралтейского района Санкт-Петербурга']").click()
        driver.find_element_by_css_selector(
            "#buildForm\:sendInstitutions_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(1) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element_by_css_selector(
            "#buildForm\:sendInstitutions_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(4) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element_by_css_selector("body.main-body").click()
        driver.find_element_by_id("buildForm:j_idt94").click()
        time.sleep(4)
        address=driver.find_element_by_id("dataForm:j_idt103").text
        driver.find_element_by_id("buildForm:exportBtn").click()
        time.sleep(4)

        rootpath = 'C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel'
        filelist = [os.path.join(rootpath, f) for f in os.listdir(rootpath)]
        filelist = [f for f in filelist if os.path.isfile(f)]
        newest = max(filelist, key=lambda x: os.stat(x).st_mtime)


        #wb = load_workbook('C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel\\Отчет о проведении исследований на COVID19 01.07.2021 - 31.07.2021.xlsx')
        wb = load_workbook(newest)
        sheet_ranges = wb['1']
        address2 = sheet_ranges['X11'].value

        assert (re.match(r'(?i)' + re.sub(r'\s', '', address) + r'$', re.sub(r'\s', '', address2)))
        #os.remove('C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel\\Отчет о проведении исследований на COVID19 01.07.2021 - 31.07.2021.xlsx')

