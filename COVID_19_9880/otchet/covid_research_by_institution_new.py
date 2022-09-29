# -*- coding: utf-8 -*-
import glob
#COV-1010 Не выгружаются адреса в отчет /reports/covid-research-by-institution

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
import logging, os
from openpyxl import load_workbook

class CovidResearchByInstitutionNew(unittest.TestCase):
    def setUp(self):
        download_dir = "C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel"
        chrome_options = webdriver.ChromeOptions()
        preferences = {"download.default_directory": download_dir,
                       "directory_upgrade": True,
                       "safebrowsing.enabled": True}
        chrome_options.add_experimental_option("prefs", preferences)
        self.driver = webdriver.Chrome("C:/Users/user/Downloads/chromedriver.exe", options=chrome_options)
        self.driver.set_window_size(1024, 600)
        self.driver.maximize_window()
        self.driver.implicitly_wait(60)
        self.verificationErrors = []
        self.accept_next_alert = True


    def test_covid_research_by_institution_new(self):
        driver = self.driver
        #driver.get("http://195.19.96.255:8981/documents/")
        driver.get("http://test.rpn19.ru/reports/reports-list")
        #driver.get("http://127.0.0.1:48080/business/dashboard/dashboard.xhtml")
        # driver.get("https://rpn19.ru:11443/documents/")
        driver.find_element(By.ID,"form:usernameInput").click()
        driver.find_element(By.ID,"form:usernameInput").clear()
        driver.find_element(By.ID,"form:usernameInput").send_keys("borisova")
        driver.find_element(By.ID,"form:passwordInput").click()
        driver.find_element(By.ID,"form:passwordInput").clear()
        driver.find_element(By.ID,"form:passwordInput").send_keys("Gi8BbtDN")
        driver.find_element(By.CSS_SELECTOR,"span.ui-button-text.ui-c").click()
        driver.find_element(By.ID,"reportsForm:j_idt79:4:j_idt81").click()

        #фильтры
        driver.find_element(By.ID,"buildForm:j_idt80_input").click()
        driver.find_element(By.ID,"buildForm:j_idt80_input").clear()
        for date in "1202.80.10":
            driver.find_element(By.ID,"buildForm:j_idt80_input").send_keys(Keys.HOME, date)
        driver.find_element(By.ID,"buildForm:j_idt82_input").click()
        driver.find_element(By.ID,"buildForm:j_idt82_input").clear()
        for date in "1202.80.61":
            driver.find_element(By.ID,"buildForm:j_idt82_input").send_keys(Keys.HOME, date)

        driver.find_element(By.ID,"buildForm:j_idt83").click()
        time.sleep(2)
        #driver.find_element(By.ID,"buildForm:j_idt79_label").click()
        #driver.find_element(By.CSS_SELECTOR,"#buildForm\:j_idt84_panel").click()
        #time.sleep(5)
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt83_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(39) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        # driver.find_element(By.CSS_SELECTOR,"#buildForm\:j_idt79_panel").click()
        #driver.find_element(By.CSS_SELECTOR,
        #    "#buildForm\:j_idt79_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(81) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()

        driver.find_element(By.ID,"buildForm:j_idt85").click()
        # driver.find_element(By.ID,"buildForm:j_idt79_label").click()
        driver.find_element(By.CSS_SELECTOR,"#buildForm\:j_idt85_panel").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(1) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(2) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(3) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(5) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(6) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()
        driver.find_element(By.CSS_SELECTOR,
            "#buildForm\:j_idt85_panel > div.ui-selectcheckboxmenu-items-wrapper > ul > li:nth-child(7) > div > div.ui-chkbox-box.ui-widget.ui-corner-all.ui-state-default > span").click()


        driver.find_element(By.CSS_SELECTOR,"body").click()
        driver.find_element(By.ID,"buildForm:j_idt95").click()
        time.sleep(2)

        assert (driver.find_element(By.ID,"dataForm:j_idt100").text == "01.08.2021 00:00 - 16.08.2021 00:00")
        assert (driver.find_element(By.ID,"dataForm:j_idt102").text == "Определение наличия РНК SARS-CoV-2")
        assert (driver.find_element(By.ID,"dataForm:j_idt104").text == "Набор реагентов для выявления РНК коронавируса SARS-CoV-2 методом ОТ-ПЦР в режиме реального времени \"РеалБест РНК SARS-CoV-2\"")
        assert (driver.find_element(By.ID,"dataForm:j_idt108").text == "11")

        driver.find_element(By.ID,"buildForm:j_idt96").click()
        time.sleep(3)

        rootpath = 'C:\\Users\\user\\PycharmProjects\\TestsCovid1\\COVID_19_9880\\otchet\\downloads_exel'
        filelist = [os.path.join(rootpath, f) for f in os.listdir(rootpath)]
        filelist = [f for f in filelist if os.path.isfile(f)]
        newest = max(filelist, key=lambda x: os.stat(x).st_mtime)

        wb = load_workbook(newest)
        #wb = newest.get_sheet_by_name('Sheet1')
        #sheet_ranges=wb.create_sheet("1")
        sheet_ranges = wb['1']
        str1 = sheet_ranges['A21'].value
        str2 = sheet_ranges['A48'].value


        assert (str1 == "Учреждение, направившее материал: АО \"АТОМПРОЕКТ\"")

if __name__ == "__main__":
    unittest.main()
